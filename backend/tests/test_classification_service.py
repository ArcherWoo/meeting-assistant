from __future__ import annotations

import os
import shutil
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

BACKEND_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
if BACKEND_ROOT not in sys.path:
    sys.path.insert(0, BACKEND_ROOT)

import services.classification_service as classification_module
from services.classification_service import classification_service
from services.retrieval_planner import RetrievalPlannerSettings
from services.storage import storage


def _pick_sample_excels(project_root: Path) -> tuple[Path, Path]:
    files = sorted((project_root / 'test_data').glob('*.xlsx'))
    assert len(files) >= 2
    return files[0], files[1]


@pytest.mark.asyncio
async def test_classification_service_generates_excel_output():
    project_root = Path(__file__).resolve().parents[2]
    template_path, data_path = _pick_sample_excels(project_root)
    output_dir = project_root / 'backend' / '.tmp-classification-test'
    if output_dir.exists():
        shutil.rmtree(output_dir, ignore_errors=True)
    output_dir.mkdir(parents=True, exist_ok=True)

    try:
        result = await classification_service.classify_excel_files(
            template_path=str(template_path.relative_to(project_root)),
            data_path=str(data_path.relative_to(project_root)),
            output_dir=output_dir,
            settings=RetrievalPlannerSettings(),
        )

        output_path = Path(result['output_path'])
        assert output_path.exists()
        assert result['processed_count'] >= 3
        assert result['matched_count'] >= 0
        assert result['review_count'] >= 0

        workbook = load_workbook(output_path, read_only=True, data_only=True)
        try:
            assert '分类结果' in workbook.sheetnames
            assert '待人工复核' in workbook.sheetnames
            assert '任务摘要' in workbook.sheetnames
        finally:
            workbook.close()
    finally:
        shutil.rmtree(output_dir, ignore_errors=True)


@pytest.mark.asyncio
async def test_classification_service_supports_import_ids(monkeypatch):
    project_root = Path(__file__).resolve().parents[2]
    template_source, data_source = _pick_sample_excels(project_root)
    imported_root = project_root / 'backend' / '.tmp-imported-files'
    template_path = imported_root / 'template-import' / template_source.name
    data_path = imported_root / 'data-import' / data_source.name
    output_dir = project_root / 'backend' / '.tmp-classification-import-test'

    if imported_root.exists():
        shutil.rmtree(imported_root, ignore_errors=True)
    if output_dir.exists():
        shutil.rmtree(output_dir, ignore_errors=True)

    template_path.parent.mkdir(parents=True, exist_ok=True)
    data_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(template_source, template_path)
    shutil.copy2(data_source, data_path)
    output_dir.mkdir(parents=True, exist_ok=True)

    async def fake_list_ppt_imports(*args, **kwargs):
        return [
            {'id': 'template-import', 'file_name': template_path.name},
            {'id': 'data-import', 'file_name': data_path.name},
        ]

    async def fake_get_ppt_import(import_id: str):
        if import_id == 'template-import':
            return {'id': import_id, 'stored_file_path': str(template_path)}
        if import_id == 'data-import':
            return {'id': import_id, 'stored_file_path': str(data_path)}
        return None

    monkeypatch.setattr(storage, 'list_ppt_imports', fake_list_ppt_imports)
    monkeypatch.setattr(storage, 'get_ppt_import', fake_get_ppt_import)
    monkeypatch.setattr(classification_module, 'IMPORTED_FILES_DIR', imported_root)

    try:
        result = await classification_service.classify_excel_files(
            template_import_id='template-import',
            data_import_id='data-import',
            output_dir=output_dir,
            settings=RetrievalPlannerSettings(),
        )
        assert Path(result['output_path']).exists()
        assert result['processed_count'] >= 3
    finally:
        shutil.rmtree(imported_root, ignore_errors=True)
        shutil.rmtree(output_dir, ignore_errors=True)

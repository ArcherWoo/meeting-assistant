from __future__ import annotations

import asyncio
import io
import json
import logging
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Literal

from openpyxl import Workbook, load_workbook

from services.knowhow_router import knowhow_router
from services.knowhow_service import knowhow_service
from services.retrieval_planner import RetrievalPlannerSettings
from services.runtime_paths import APP_HOME, CLASSIFICATION_OUTPUTS_DIR, DATA_DIR, IMPORTED_FILES_DIR, PROJECT_ROOT
from services.storage import storage
from utils.text_utils import extract_han_segments

logger = logging.getLogger(__name__)


ClassificationMode = Literal["strict", "balanced", "recall"]


@dataclass(frozen=True)
class TaxonomyPath:
    path_id: str
    levels: tuple[str, str, str, str, str]
    full_path: str
    searchable_text: str
    keywords: tuple[str, ...]


@dataclass(frozen=True)
class RowClassification:
    item_name: str
    matched: bool
    path_id: str | None
    levels: tuple[str, str, str, str, str]
    full_path: str
    confidence: float
    reason: str
    matched_rule_ids: tuple[str, ...]
    matched_rule_titles: tuple[str, ...]
    needs_review: bool
    strategy: str


class ClassificationService:
    def __init__(self) -> None:
        self._allowed_roots = tuple(root.resolve() for root in (PROJECT_ROOT, APP_HOME, DATA_DIR))

    async def classify_excel_files(
        self,
        *,
        template_path: str = "",
        data_path: str = "",
        template_import_id: str = "",
        data_import_id: str = "",
        template_sheet: str = "",
        data_sheet: str = "",
        name_column: int = 2,
        knowhow_categories: list[str] | None = None,
        mode: ClassificationMode = "balanced",
        review_threshold: float = 0.55,
        settings: RetrievalPlannerSettings | None = None,
        user_id: str | None = None,
        group_id: str | None = None,
        is_admin: bool = False,
        output_dir: Path | None = None,
    ) -> dict[str, Any]:
        resolved_template_path = await self._resolve_source_file(
            raw_path=template_path,
            import_id=template_import_id,
            user_id=user_id,
            group_id=group_id,
            is_admin=is_admin,
        )
        resolved_data_path = await self._resolve_source_file(
            raw_path=data_path,
            import_id=data_import_id,
            user_id=user_id,
            group_id=group_id,
            is_admin=is_admin,
        )
        output_root = output_dir or self._build_user_output_dir(user_id)
        output_root.mkdir(parents=True, exist_ok=True)

        taxonomy_paths = await asyncio.to_thread(
            self._parse_template_workbook,
            resolved_template_path,
            template_sheet,
        )
        if not taxonomy_paths:
            raise ValueError("分类模板中未解析出有效的分类路径")

        try:
            knowhow_rules = await knowhow_service.list_rules(
                active_only=True,
                user_id=user_id,
                group_id=group_id,
                is_admin=is_admin,
            )
            category_profiles = await knowhow_service.list_categories(
                user_id=user_id,
                group_id=group_id,
                is_admin=is_admin,
            )
        except RuntimeError:
            knowhow_rules = []
            category_profiles = []
        if knowhow_categories:
            normalized_categories = {
                self._normalize_text(item)
                for item in knowhow_categories
                if self._normalize_text(item)
            }
            knowhow_rules = [
                rule for rule in knowhow_rules
                if self._normalize_text(rule.get("category")) in normalized_categories
            ]
        if knowhow_categories:
            normalized_categories = {
                self._normalize_text(item)
                for item in knowhow_categories
                if self._normalize_text(item)
            }
            category_profiles = [
                category for category in category_profiles
                if self._normalize_text(category.get("name")) in normalized_categories
            ]

        workbook_payload = await asyncio.to_thread(
            self._parse_data_workbook,
            resolved_data_path,
            data_sheet,
            name_column,
        )
        if not workbook_payload["rows"]:
            raise ValueError("待分类数据中没有可处理的有效行")

        unique_names = list(dict.fromkeys(row["item_name"] for row in workbook_payload["rows"] if row["item_name"]))
        decisions = await self._classify_unique_names(
            names=unique_names,
            taxonomy_paths=taxonomy_paths,
            knowhow_rules=knowhow_rules,
            category_profiles=category_profiles,
            mode=mode,
            review_threshold=review_threshold,
            settings=settings,
        )

        processed_rows: list[dict[str, Any]] = []
        review_rows: list[dict[str, Any]] = []
        rule_hits: list[str] = []
        for row in workbook_payload["rows"]:
            decision = decisions.get(row["item_name"]) or self._build_empty_result(
                item_name=row["item_name"],
                reason="未找到分类结果",
            )
            row_payload = {
                **row,
                "matched": "是" if decision.matched else "否",
                "一级分类": decision.levels[0],
                "二级分类": decision.levels[1],
                "三级分类": decision.levels[2],
                "四级分类": decision.levels[3],
                "五级分类": decision.levels[4],
                "匹配路径": decision.full_path,
                "置信度": round(decision.confidence, 3),
                "命中规则": "；".join(decision.matched_rule_titles),
                "匹配原因": decision.reason,
                "需人工复核": "是" if decision.needs_review else "否",
                "判定策略": decision.strategy,
            }
            processed_rows.append(row_payload)
            rule_hits.extend(decision.matched_rule_ids)
            if decision.needs_review:
                review_rows.append(row_payload)

        if rule_hits:
            try:
                await knowhow_service.record_rule_hits(list(dict.fromkeys(rule_hits)))
            except RuntimeError:
                logger.debug("skip knowhow hit recording because storage is not initialized")

        output_filename = self._build_output_filename(resolved_data_path)
        output_path = output_root / output_filename
        await asyncio.to_thread(
            self._write_output_workbook,
            output_path,
            workbook_payload["headers"],
            processed_rows,
            review_rows,
            {
                "template_file": str(resolved_template_path),
                "data_file": str(resolved_data_path),
                "template_sheet": template_sheet or workbook_payload["template_sheet_hint"],
                "data_sheet": workbook_payload["sheet_name"],
                "name_column": name_column,
                "mode": mode,
                "review_threshold": review_threshold,
                "template_path_count": len(taxonomy_paths),
                "knowhow_rule_count": len(knowhow_rules),
                "processed_count": len(processed_rows),
                "matched_count": sum(1 for item in processed_rows if item["matched"] == "是"),
                "review_count": len(review_rows),
            },
        )

        matched_count = sum(1 for item in processed_rows if item["matched"] == "是")
        review_count = len(review_rows)
        summary = (
            f"已处理 {len(processed_rows)} 条数据，命中分类 {matched_count} 条，"
            f"待人工复核 {review_count} 条，结果文件已生成。"
        )
        preview_rows = processed_rows[:5]
        return {
            "summary": summary,
            "output_path": str(output_path),
            "output_filename": output_filename,
            "processed_count": len(processed_rows),
            "matched_count": matched_count,
            "review_count": review_count,
            "template_path_count": len(taxonomy_paths),
            "knowhow_rule_count": len(knowhow_rules),
            "preview_rows": preview_rows,
            "structured_payload": {
                "summary": summary,
                "template_file": str(resolved_template_path),
                "data_file": str(resolved_data_path),
                "output_file": str(output_path),
                "processed_count": len(processed_rows),
                "matched_count": matched_count,
                "review_count": review_count,
                "mode": mode,
                "review_threshold": review_threshold,
                "preview_rows": preview_rows,
            },
        }

    async def _classify_unique_names(
        self,
        *,
        names: list[str],
        taxonomy_paths: list[TaxonomyPath],
        knowhow_rules: list[dict[str, Any]],
        category_profiles: list[dict[str, Any]],
        mode: ClassificationMode,
        review_threshold: float,
        settings: RetrievalPlannerSettings | None,
    ) -> dict[str, RowClassification]:
        semaphore = asyncio.Semaphore(3)

        async def _classify(name: str) -> tuple[str, RowClassification]:
            async with semaphore:
                return name, await self._classify_one(
                    item_name=name,
                    taxonomy_paths=taxonomy_paths,
                    knowhow_rules=knowhow_rules,
                    category_profiles=category_profiles,
                    mode=mode,
                    review_threshold=review_threshold,
                    settings=settings,
                )

        pairs = await asyncio.gather(*[_classify(name) for name in names])
        return dict(pairs)

    async def _classify_one(
        self,
        *,
        item_name: str,
        taxonomy_paths: list[TaxonomyPath],
        knowhow_rules: list[dict[str, Any]],
        category_profiles: list[dict[str, Any]],
        mode: ClassificationMode,
        review_threshold: float,
        settings: RetrievalPlannerSettings | None,
    ) -> RowClassification:
        candidates = self._select_candidates(item_name, taxonomy_paths, limit=8 if mode == "recall" else 6)
        if not candidates:
            return self._build_empty_result(item_name=item_name, reason="未找到可用候选分类")

        top_score = candidates[0]["score"]
        second_score = candidates[1]["score"] if len(candidates) > 1 else 0.0
        matched_rules: list[dict[str, Any]] = []
        if knowhow_rules:
            routing = await knowhow_router.retrieve_rules(
                item_name,
                knowhow_rules,
                category_profiles=category_profiles,
                limit=4,
                settings=settings,
            )
            matched_rules = list(routing.rules)

        heuristic_choice = self._apply_heuristic_choice(
            item_name=item_name,
            candidates=candidates,
            mode=mode,
            review_threshold=review_threshold,
            top_score=top_score,
            second_score=second_score,
            matched_rules=matched_rules,
        )
        if heuristic_choice is not None:
            return heuristic_choice

        llm_choice = await self._classify_with_llm(
            item_name=item_name,
            candidates=candidates,
            matched_rules=matched_rules,
            review_threshold=review_threshold,
            settings=settings,
        )
        if llm_choice is not None:
            return llm_choice

        top_candidate = candidates[0]["path"]
        confidence = self._score_to_confidence(top_score)
        return RowClassification(
            item_name=item_name,
            matched=confidence >= 0.35,
            path_id=top_candidate.path_id if confidence >= 0.35 else None,
            levels=top_candidate.levels if confidence >= 0.35 else ("", "", "", "", ""),
            full_path=top_candidate.full_path if confidence >= 0.35 else "",
            confidence=confidence,
            reason="未能获取稳定的 LLM 判定，已回退为候选相似度最高的路径",
            matched_rule_ids=tuple(str(rule.get("id") or "") for rule in matched_rules if str(rule.get("id") or "").strip()),
            matched_rule_titles=tuple(
                str(rule.get("title") or "未命名规则").strip()
                for rule in matched_rules
                if str(rule.get("id") or "").strip()
            ),
            needs_review=True,
            strategy="fallback_top_candidate",
        )

    async def _classify_with_llm(
        self,
        *,
        item_name: str,
        candidates: list[dict[str, Any]],
        matched_rules: list[dict[str, Any]],
        review_threshold: float,
        settings: RetrievalPlannerSettings | None,
    ) -> RowClassification | None:
        if not settings or not settings.is_configured:
            return None

        prompt = self._build_llm_prompt(item_name=item_name, candidates=candidates, matched_rules=matched_rules)
        try:
            response = await knowhow_router._llm_service.chat(
                messages=[
                    {
                        "role": "system",
                        "content": (
                            "你是分类执行器。请只在给定候选路径中选择最合适的一条，或明确返回 no_match。"
                            "必须输出 JSON，不要输出额外说明。"
                        ),
                    },
                    {"role": "user", "content": prompt},
                ],
                model=settings.model,
                temperature=0.1,
                max_tokens=420,
                api_url=settings.api_url,
                api_key=settings.api_key,
                user_id=settings.user_id,
                request_kind="lightweight",
            )
        except Exception:
            logger.debug("classification llm failed for item=%s", item_name, exc_info=True)
            return None

        payload = self._extract_json_payload(response)
        if not payload:
            return None

        chosen_path_id = self._normalize_text(payload.get("path_id"))
        confidence = self._normalize_confidence(payload.get("confidence"), default=0.35)
        reason = self._normalize_text(payload.get("reason")) or "LLM 已在候选路径中完成判定"
        matched_rule_ids = [
            str(rule_id).strip()
            for rule_id in payload.get("matched_rule_ids", [])
            if str(rule_id).strip()
        ]
        matched_rule_id_set = set(matched_rule_ids)
        matched_rule_titles = tuple(
            str(rule.get("title") or "未命名规则").strip()
            for rule in matched_rules
            if str(rule.get("id") or "").strip() in matched_rule_id_set
        )
        candidate_map = {candidate["path"].path_id: candidate["path"] for candidate in candidates}
        if chosen_path_id.lower() in {"no_match", "none", "null", ""} or chosen_path_id not in candidate_map:
            return self._build_empty_result(
                item_name=item_name,
                reason=reason,
                confidence=confidence,
                matched_rule_ids=tuple(matched_rule_ids),
                matched_rule_titles=matched_rule_titles,
                strategy="llm_no_match",
            )

        selected_path = candidate_map[chosen_path_id]
        needs_review = bool(payload.get("needs_review")) or confidence < review_threshold
        return RowClassification(
            item_name=item_name,
            matched=True,
            path_id=selected_path.path_id,
            levels=selected_path.levels,
            full_path=selected_path.full_path,
            confidence=confidence,
            reason=reason,
            matched_rule_ids=tuple(matched_rule_ids),
            matched_rule_titles=matched_rule_titles,
            needs_review=needs_review,
            strategy="llm_candidate_selection",
        )

    def _build_llm_prompt(
        self,
        *,
        item_name: str,
        candidates: list[dict[str, Any]],
        matched_rules: list[dict[str, Any]],
    ) -> str:
        candidate_lines = [
            f"- {item['path'].path_id}: {item['path'].full_path}（候选分数 {item['score']:.2f}）"
            for item in candidates[:6]
        ]
        if matched_rules:
            rule_lines = [
                f"- {str(rule.get('id') or '').strip()}: {str(rule.get('title') or '未命名规则').strip()} - {str(rule.get('rule_text') or '').strip()}"
                for rule in matched_rules[:4]
            ]
            rules_text = "\n".join(rule_lines)
        else:
            rules_text = "- 当前没有命中的业务规则"

        return (
            f"待分类名称：{item_name}\n\n"
            "候选分类路径如下：\n"
            f"{chr(10).join(candidate_lines)}\n\n"
            "可参考的业务规则：\n"
            f"{rules_text}\n\n"
            "请返回 JSON：\n"
            "{\n"
            '  "path_id": "候选 path_id 或 no_match",\n'
            '  "confidence": 0.0,\n'
            '  "reason": "简短原因",\n'
            '  "matched_rule_ids": ["命中的规则 id"],\n'
            '  "needs_review": false\n'
            "}\n"
            "如果名称本身并不属于这些候选路径，也不要硬选，直接返回 no_match。"
        )

    def _apply_heuristic_choice(
        self,
        *,
        item_name: str,
        candidates: list[dict[str, Any]],
        mode: ClassificationMode,
        review_threshold: float,
        top_score: float,
        second_score: float,
        matched_rules: list[dict[str, Any]],
    ) -> RowClassification | None:
        top_candidate = candidates[0]["path"]
        item_normalized = self._normalize_text(item_name).lower()
        leaf = self._normalize_text(top_candidate.levels[-1]).lower()
        direct_leaf_hit = bool(leaf and leaf in item_normalized)
        strong_gap = top_score >= 7.5 and (top_score - second_score >= 2.0)
        if not direct_leaf_hit and mode == "strict":
            return None

        if direct_leaf_hit and strong_gap:
            confidence = max(self._score_to_confidence(top_score), 0.82)
            return RowClassification(
                item_name=item_name,
                matched=True,
                path_id=top_candidate.path_id,
                levels=top_candidate.levels,
                full_path=top_candidate.full_path,
                confidence=confidence,
                reason="名称与候选叶子分类直接匹配，且与次优候选差距明显",
                matched_rule_ids=tuple(str(rule.get("id") or "") for rule in matched_rules if str(rule.get("id") or "").strip()),
                matched_rule_titles=tuple(
                    str(rule.get("title") or "未命名规则").strip()
                    for rule in matched_rules
                    if str(rule.get("id") or "").strip()
                ),
                needs_review=confidence < review_threshold,
                strategy="heuristic_direct_leaf",
            )
        return None

    def _select_candidates(self, item_name: str, taxonomy_paths: list[TaxonomyPath], *, limit: int) -> list[dict[str, Any]]:
        scored: list[dict[str, Any]] = []
        query = self._normalize_text(item_name)
        query_lower = query.lower()
        query_terms = self._extract_terms(query)
        for path in taxonomy_paths:
            score = 0.0
            leaf = self._normalize_text(path.levels[-1]).lower()
            if leaf and leaf in query_lower:
                score += 6.5
            for level in path.levels:
                level_text = self._normalize_text(level).lower()
                if level_text and level_text in query_lower:
                    score += 1.8
            for term in query_terms:
                if term and term in path.searchable_text:
                    score += 1.2
            shared_terms = sum(1 for keyword in path.keywords if keyword in query_lower)
            score += shared_terms * 0.5
            if score > 0:
                scored.append({"path": path, "score": score})

        if not scored:
            scored = [{"path": path, "score": 0.0} for path in taxonomy_paths[:limit]]

        scored.sort(
            key=lambda item: (
                item["score"],
                len(item["path"].full_path),
            ),
            reverse=True,
        )
        return scored[:limit]

    def _parse_template_workbook(self, file_path: Path, sheet_name: str) -> list[TaxonomyPath]:
        workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
        try:
            sheet = workbook[sheet_name] if sheet_name and sheet_name in workbook.sheetnames else workbook[workbook.sheetnames[0]]
            last_values = ["", "", "", "", ""]
            paths: list[TaxonomyPath] = []
            seen_paths: set[str] = set()
            for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                values = [self._normalize_text(row[col_index] if col_index < len(row) else "") for col_index in range(5)]
                if row_index == 1:
                    continue
                if not any(values):
                    continue
                for index, value in enumerate(values):
                    if value:
                        last_values[index] = value
                        for clear_index in range(index + 1, 5):
                            last_values[clear_index] = ""
                levels = tuple(last_values)
                full_path = " > ".join(level for level in levels if level)
                if not full_path or full_path in seen_paths:
                    continue
                seen_paths.add(full_path)
                path_id = f"path-{len(paths) + 1}"
                searchable_parts = [full_path, " ".join(level for level in levels if level)]
                keywords = self._extract_terms(" ".join(level for level in levels if level))
                paths.append(
                    TaxonomyPath(
                        path_id=path_id,
                        levels=levels,
                        full_path=full_path,
                        searchable_text=" ".join(part.lower() for part in searchable_parts if part).strip(),
                        keywords=tuple(keywords),
                    )
                )
            return paths
        finally:
            workbook.close()

    def _parse_data_workbook(self, file_path: Path, sheet_name: str, name_column: int) -> dict[str, Any]:
        workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
        try:
            sheet = workbook[sheet_name] if sheet_name and sheet_name in workbook.sheetnames else workbook[workbook.sheetnames[0]]
            headers: list[str] = []
            rows: list[dict[str, Any]] = []
            for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                values = list(row)
                if row_index == 1:
                    headers = [
                        self._normalize_text(cell) or f"列{column_index + 1}"
                        for column_index, cell in enumerate(values)
                    ]
                    continue
                if not any(value is not None and self._normalize_text(value) for value in values):
                    continue
                name_value = self._normalize_text(values[name_column - 1] if len(values) >= name_column else "")
                original_row = {
                    headers[column_index] if column_index < len(headers) else f"列{column_index + 1}": value
                    for column_index, value in enumerate(values)
                }
                original_row["item_name"] = name_value
                original_row["row_index"] = row_index
                rows.append(original_row)
            return {
                "headers": headers,
                "rows": rows,
                "sheet_name": sheet.title,
                "template_sheet_hint": "",
            }
        finally:
            workbook.close()

    def _write_output_workbook(
        self,
        output_path: Path,
        original_headers: list[str],
        processed_rows: list[dict[str, Any]],
        review_rows: list[dict[str, Any]],
        summary_payload: dict[str, Any],
    ) -> None:
        workbook = Workbook()
        result_sheet = workbook.active
        result_sheet.title = "分类结果"

        result_headers = [
            *original_headers,
            "一级分类",
            "二级分类",
            "三级分类",
            "四级分类",
            "五级分类",
            "匹配路径",
            "置信度",
            "命中规则",
            "匹配原因",
            "需人工复核",
            "判定策略",
        ]
        result_sheet.append(result_headers)
        for row in processed_rows:
            result_sheet.append([row.get(header, "") for header in result_headers])

        review_sheet = workbook.create_sheet("待人工复核")
        review_sheet.append(result_headers)
        for row in review_rows:
            review_sheet.append([row.get(header, "") for header in result_headers])

        summary_sheet = workbook.create_sheet("任务摘要")
        summary_sheet.append(["字段", "值"])
        for key, value in summary_payload.items():
            summary_sheet.append([key, json.dumps(value, ensure_ascii=False) if isinstance(value, (dict, list)) else value])

        workbook.save(output_path)

    async def _resolve_source_file(
        self,
        *,
        raw_path: str,
        import_id: str,
        user_id: str | None,
        group_id: str | None,
        is_admin: bool,
    ) -> Path:
        normalized_import_id = self._normalize_text(import_id)
        if normalized_import_id:
            return await self._resolve_import_file(
                normalized_import_id,
                user_id=user_id,
                group_id=group_id,
                is_admin=is_admin,
            )
        return self._resolve_input_path(raw_path)

    async def _resolve_import_file(
        self,
        import_id: str,
        *,
        user_id: str | None,
        group_id: str | None,
        is_admin: bool,
    ) -> Path:
        accessible_imports = await storage.list_ppt_imports(
            user_id=user_id,
            group_id=group_id,
            is_admin=is_admin,
        )
        if not any(str(item.get("id") or "").strip() == import_id for item in accessible_imports):
            raise PermissionError("当前用户无权使用该导入文件")

        import_row = await storage.get_ppt_import(import_id)
        if not import_row:
            raise FileNotFoundError("导入文件不存在")

        stored_file_path = self._normalize_text(import_row.get("stored_file_path"))
        if not stored_file_path:
            raise FileNotFoundError("该导入文件未保留原始工作簿，暂时无法用于 Excel 分类")

        candidate = Path(stored_file_path).resolve()
        if not candidate.is_relative_to(IMPORTED_FILES_DIR.resolve()):
            raise ValueError("导入文件不在允许的应用数据目录下")
        if not candidate.exists() or not candidate.is_file():
            raise FileNotFoundError(f"导入文件不存在：{candidate}")
        return candidate

    def _resolve_input_path(self, raw_path: str) -> Path:
        normalized = self._normalize_text(raw_path)
        if not normalized:
            raise ValueError("文件路径不能为空")
        candidate = Path(normalized)
        if not candidate.is_absolute():
            candidate = (PROJECT_ROOT / candidate).resolve()
        else:
            candidate = candidate.resolve()
        if not any(candidate.is_relative_to(root) for root in self._allowed_roots):
            raise ValueError("当前仅允许访问项目目录或应用数据目录中的文件")
        if not candidate.exists() or not candidate.is_file():
            raise FileNotFoundError(f"文件不存在：{candidate}")
        return candidate

    def _build_user_output_dir(self, user_id: str | None) -> Path:
        normalized_user_id = self._normalize_text(user_id)
        safe_segment = re.sub(r"[^a-zA-Z0-9._-]+", "_", normalized_user_id) if normalized_user_id else "anonymous"
        return (CLASSIFICATION_OUTPUTS_DIR / safe_segment).resolve()

    @staticmethod
    def _build_output_filename(data_path: Path) -> str:
        stem = data_path.stem
        timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
        return f"{stem}-分类结果-{timestamp}.xlsx"

    @staticmethod
    def _normalize_text(value: Any) -> str:
        return str(value or "").strip()

    def _extract_terms(self, text: str) -> list[str]:
        normalized = self._normalize_text(text)
        if not normalized:
            return []
        english_terms = re.findall(r"[a-zA-Z0-9][a-zA-Z0-9\\-_/]{1,}", normalized.lower())
        han_terms = extract_han_segments(normalized, min_length=1, max_length=4)
        merged: list[str] = []
        seen: set[str] = set()
        for item in [normalized.lower(), *english_terms, *han_terms]:
            cleaned = self._normalize_text(item).lower()
            if not cleaned or cleaned in seen:
                continue
            seen.add(cleaned)
            merged.append(cleaned)
        return merged

    @staticmethod
    def _normalize_confidence(value: Any, *, default: float) -> float:
        try:
            confidence = float(value)
        except (TypeError, ValueError):
            confidence = default
        return max(0.0, min(1.0, confidence))

    @staticmethod
    def _score_to_confidence(score: float) -> float:
        if score >= 8:
            return 0.92
        if score >= 6:
            return 0.8
        if score >= 4:
            return 0.62
        if score >= 2:
            return 0.42
        return 0.2

    @staticmethod
    def _extract_json_payload(response: dict[str, Any]) -> dict[str, Any] | None:
        content = (
            response.get("choices", [{}])[0]
            .get("message", {})
            .get("content", "")
        )
        if not isinstance(content, str) or not content.strip():
            return None
        stripped = content.strip()
        try:
            return json.loads(stripped)
        except json.JSONDecodeError:
            match = re.search(r"\{.*\}", stripped, re.DOTALL)
            if not match:
                return None
            try:
                return json.loads(match.group(0))
            except json.JSONDecodeError:
                return None

    @staticmethod
    def _build_empty_result(
        *,
        item_name: str,
        reason: str,
        confidence: float = 0.0,
        matched_rule_ids: tuple[str, ...] = (),
        matched_rule_titles: tuple[str, ...] = (),
        strategy: str = "no_match",
    ) -> RowClassification:
        return RowClassification(
            item_name=item_name,
            matched=False,
            path_id=None,
            levels=("", "", "", "", ""),
            full_path="",
            confidence=confidence,
            reason=reason,
            matched_rule_ids=matched_rule_ids,
            matched_rule_titles=matched_rule_titles,
            needs_review=True,
            strategy=strategy,
        )


classification_service = ClassificationService()
